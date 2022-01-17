from openpyxl import load_workbook
from openpyxl import Workbook
load_wb = load_workbook("C:/Users/tofan/Documents/카카오톡 받은 파일/합친파일.xlsx", data_only=True)

# 시트 이름으로 불러오기
load_ws = load_wb['Sheet1']

# 모든 행과 열 데이터 출력
all_values = []

for row in load_ws.rows:
    row_value = []
    for (index, cell) in enumerate(row):
        if index == 2:
            # print(cell.value)
            cells = str(cell.value).split(",")
            result = []
            for keyword in cells:
                # print("현재 키워드는 : ", keyword)
                if keyword\
                        .replace(" ", "")\
                        .replace("-", "")\
                        .replace("/", "")\
                        .replace("(", "")\
                        .replace(")", "")\
                        .replace("’", "")\
                        .replace('"', "")\
                        .replace('`', "")\
                        .replace('.', "")\
                        .replace("\\", "")\
                        .replace("'", "")\
                        .encode().isalnum() or "http://" in keyword:
                    # print("영어당")
                    continue
                else:
                    if len(keyword) > 0:
                        result.append(keyword.replace("’", ""))
            # 끝나면 cell.value를 새롭게 변경한다.
            result_str = ""
            for (idx, keyword) in enumerate(result):
                result_str += keyword
                # 마지막에는 콤마와 공백을 넣어선 안된다.
                if idx != len(result) - 1:
                    result_str += ", "
            cell.value = result_str
        row_value.append(cell.value)
    all_values.append(row_value)

# 이제 다시 파일 쓰자.
write_wb = Workbook()
write_ws = write_wb.active # Sheet1에 입력

# 행단위로 추가
for v in all_values:
    write_ws.append(v)

write_wb.save("C:/Users/tofan/Documents/카카오톡 받은 파일/합친파일_수정본.xlsx")